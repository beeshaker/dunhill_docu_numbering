from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

from .branding import create_faded_logo


def _add_logo_watermark(ws, watermark_path: Path) -> None:
    try:
        img = XLImage(str(watermark_path))
        img.anchor = "D4"
        ws.add_image(img)
    except Exception:
        pass


def insert_reference_excel(
    input_path: Path,
    output_path: Path,
    reference_number: str,
    *,
    add_watermark: bool = False,
    logo_path: Path | None = None,
) -> Path:
    wb = load_workbook(str(input_path))
    ws = wb.active

    ws.insert_rows(1)
    cell = ws.cell(row=1, column=1)
    cell.value = f"Ref No.: {reference_number}"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")

    max_col = max(ws.max_column, 4)
    try:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(max_col, 6))
    except Exception:
        pass

    if add_watermark:
        watermark_path = create_faded_logo(logo_path, opacity=0.10, max_width_px=600)
        if watermark_path:
            _add_logo_watermark(ws, watermark_path)

    wb.save(str(output_path))
    return output_path
